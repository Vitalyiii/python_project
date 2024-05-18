#импорт нужных библиотек

import tkinter as tk
import sys
import os
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import ttk
from tkinter import messagebox

# load_data загружает файл
def load_data():
    path = 'document.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    print(list_values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)
#внесение данных и отображение в программе
def insert_row():
    years = year.get()
    indicators_one = index_one.get()
    indicators_two = index_two.get()

    print(years, indicators_one, indicators_two)

    path = 'document.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [years, indicators_one, indicators_two]
    sheet.append(row_values)
    workbook.save(path)

    treeview.insert('', tk.END, values=row_values)

    year.delete(0, 'end')
    year.insert(0, 'Год')
    index_one.delete(0, 'end')
    index_one.insert(0, 'Показатель №1 %')
    index_two.delete(0, 'end')
    index_two.insert(0, 'Показатель №2 %')
#удаление последней строки с перезапуском программы
def delete_last_row():
    path = 'document.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    last_row = sheet.max_row
    sheet.delete_rows(last_row, 1)
    workbook.save(path)
    messagebox.showinfo(title="Уведомление", message="Крайняя строка удалена. Программа перезапустится")
    os.execl(sys.executable, sys.executable, *sys.argv)
# анализ соотношения двух показателей
def analysis():
    df = pd.read_excel('document.xlsx')

    year = df['Год'].astype(str).tolist()
    index_one = df['Показатель №1 %'].tolist()
    index_two = df['Показатель №2 %'].tolist()

    width = 0.35

    fig, ax = plt.subplots()

    ax.bar(year, index_one, width, label='Показатель №1')
    ax.bar(year, index_two, width, bottom=index_one, label='Показатель №2')
    ax.set_ylabel('Соотношение, в %')
    ax.set_title('Соотношение двух показателей ')
    ax.legend(loc='lower left', title='Анализ')
    plt.show()

root = tk.Tk()

root.title("Документ")
root.geometry("700x280")
root.resizable(False, False)

frame = ttk.Frame(root)
frame.pack()
# создаем виджет
widgets_frame = ttk.LabelFrame(frame, text='Ввод данных:')
widgets_frame.grid(row=0, column=0, padx=20, pady=6)

year = ttk.Entry(widgets_frame)
year.insert(0, 'Год')
year.bind('<FocusIn>', lambda e: year.delete('0', 'e'))
year.grid(row=0, column=0, sticky='ew', padx=5, pady=(0.5))

index_one = ttk.Entry(widgets_frame)
index_one.insert(0, 'Показатель №1 %')
index_one.bind('<FocusIn>', lambda e: index_one.delete('0', 'e'))
index_one.grid(row=1, column=0, sticky='ew', padx=5, pady=5)

index_two = ttk.Entry(widgets_frame)
index_two.insert(0, 'Показатель №2 %')
index_two.bind('<FocusIn>', lambda e: index_two.delete('0', 'e'))
index_two.grid(row=2, column=0, sticky='ew', padx=5, pady=(0.5))
# создаем кнопки для вызова функций
button = ttk.Button(widgets_frame, text='Внести данные', command=insert_row)
button.grid(row=3, column=0, sticky='nsew', padx=5, pady=2)

button1 = ttk.Button(widgets_frame, text='Удалить посл. строку',command=delete_last_row)
button1.grid(row=4, column=0, sticky='nsew', padx=5, pady=2)

button2 = ttk.Button(widgets_frame, text='Анализ',command=analysis)
button2.grid(row=5, column=0, sticky='nsew', padx=5, pady=2)

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)

treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side='right', fill='y')
#имена столбцов и все что связано с отображением
cols = ("Год", "Показатель №1 %", "Показатель №2 %")

treeview = ttk.Treeview(treeFrame, show='headings', yscrollcommand=treeScroll.set, columns=cols, height=13)

treeview.column('Год', width=130)
treeview.column('Показатель №1 %', width=150)
treeview.column('Показатель №1 %', width=150)

treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()

root.mainloop()
